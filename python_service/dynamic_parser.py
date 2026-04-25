#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
军事指标动态解析器
完全根据Excel列结构动态识别，不写死任何层级名称

Excel格式：
- 第1列: 层级（含描述信息，可合并单元格）
- 第2列: 一级维度
- 第3列: 二级维度（最小粒度）
"""

import sys
import json
import re
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from collections import defaultdict
import pandas as pd
import numpy as np

# ============================================================================
# 数据结构定义
# ============================================================================

@dataclass
class LevelInfo:
    """层级信息"""
    name: str                    # 层级名称
    description: str              # 层级描述（对应级别、具体职级、评估重点）
    primary_dimensions: List['PrimaryDimension'] = field(default_factory=list)
    sort_order: int = 0         # 排序顺序


@dataclass
class PrimaryDimension:
    """一级维度"""
    name: str                    # 一级维度名称
    code: str                    # 编码（自动生成）
    sort_order: int = 0         # 排序顺序
    secondary_dimensions: List['SecondaryDimension'] = field(default_factory=list)


@dataclass
class SecondaryDimension:
    """二级维度（最小粒度）"""
    name: str                    # 二级维度名称
    code: str                    # 编码（自动生成）
    sort_order: int = 0          # 排序顺序
    metric_type: str = 'QUALITATIVE'  # 指标性质：定性/定量


@dataclass
class ParsedResult:
    """解析结果"""
    levels: List[LevelInfo]
    statistics: Dict

# ============================================================================
# 动态解析器
# ============================================================================

class DynamicIndicatorParser:
    """
    动态指标Excel解析器
    
    核心逻辑：
    1. 按列位置读取（第1列=层级，第2列=一级维度，第3列=二级维度）
    2. 自动检测合并单元格（用于层级描述）
    3. 不写死任何层级名称，完全动态识别
    """
    
    def __init__(self):
        self.levels: List[LevelInfo] = []
        self.merged_cells: Dict = {}  # 存储合并单元格信息
        self.statistics: Dict = {}    # 存储统计信息
        
    def parse_file(self, file_path: str) -> ParsedResult:
        """从文件解析"""
        # 读取Excel（不处理合并单元格）
        df = pd.read_excel(file_path, header=None, engine='openpyxl')
        
        # 读取合并单元格信息
        self._read_merged_cells(file_path)
        
        return self._parse_dataframe(df)
    
    def parse_bytes(self, file_bytes: bytes) -> ParsedResult:
        """从字节流解析（用于接收上传的文件）"""
        import io
        
        # 读取Excel
        df = pd.read_excel(io.BytesIO(file_bytes), header=None, engine='openpyxl')
        
        # 读取合并单元格信息
        self._read_merged_cells_from_bytes(file_bytes)
        
        return self._parse_dataframe(df)
    
    def _read_merged_cells(self, file_path: str):
        """读取合并单元格信息"""
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        self.merged_cells = {}
        for merged_range in ws.merged_cells.ranges:
            # 合并范围，如 "A1:A5"
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            
            # 只处理第一列的合并单元格（层级列）
            if min_col == 1:
                cell_value = ws.cell(min_row, min_col).value
                if cell_value:
                    # 存储合并的行范围
                    self.merged_cells[(min_row, max_row)] = str(cell_value)
    
    def _read_merged_cells_from_bytes(self, file_bytes: bytes):
        """从字节流读取合并单元格信息"""
        import io
        from openpyxl import load_workbook
        
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        
        self.merged_cells = {}
        for merged_range in ws.merged_cells.ranges:
            min_col = merged_range.min_col
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            
            if min_col == 1:
                cell_value = ws.cell(min_row, min_col).value
                if cell_value:
                    self.merged_cells[(min_row, max_row)] = str(cell_value)
    
    def _parse_dataframe(self, df: pd.DataFrame) -> ParsedResult:
        """核心解析逻辑"""
        self.levels = []

        current_level: Optional[LevelInfo] = None
        current_primary: Optional[PrimaryDimension] = None
        primary_counter = 0
        secondary_counter = 0
        level_counter = 0

        # 遍历每一行
        for row_idx in range(len(df)):
            # 获取四列的值
            col0_val = self._get_cell_value(df, row_idx, 0)  # 层级
            col1_val = self._get_cell_value(df, row_idx, 1)  # 一级维度
            col2_val = self._get_cell_value(df, row_idx, 2)  # 二级维度
            col3_val = self._get_cell_value(df, row_idx, 3)  # 指标性质（定性/定量）

            # 跳过空行（全部为空）
            if not col0_val and not col1_val and not col2_val:
                continue

            # 跳过标题行（如果第一行是表头的话）
            # 简单检查：如果第一列值是"层级"或类似表头名称
            if row_idx == 0 and col0_val in ['层级', '级别', 'level', 'Level']:
                continue

            # 检测是否是层级标题行
            if self._is_level_row(df, col0_val, row_idx):
                # 保存之前的层级
                if current_level:
                    self.levels.append(current_level)

                # 创建新层级
                level_counter += 1
                level_name = self._extract_level_name(col0_val)
                level_desc = self._extract_level_description(col0_val)

                current_level = LevelInfo(
                    name=level_name,
                    description=level_desc,
                    sort_order=level_counter
                )
                current_primary = None

                # 如果该行有一级维度，继续处理
                if col1_val and col1_val.strip():
                    # 检查一级维度是否已存在（防止重复）
                    existing_primary = None
                    for p in current_level.primary_dimensions:
                        if p.name == col1_val:
                            existing_primary = p
                            break
                    if existing_primary:
                        current_primary = existing_primary
                    else:
                        primary_counter += 1
                        primary_code = self._generate_code(col1_val)
                        current_primary = PrimaryDimension(
                            name=col1_val,
                            code=primary_code,
                            sort_order=primary_counter
                        )
                        current_level.primary_dimensions.append(current_primary)

                    # 如果有二级维度
                    if col2_val and col2_val.strip():
                        # 解析指标性质
                        metric_type = self._parse_metric_type(col3_val)
                        # 检查二级维度是否已存在
                        existing_sec = None
                        for s in current_primary.secondary_dimensions:
                            if s.name == col2_val:
                                existing_sec = s
                                break
                        if not existing_sec:
                            secondary_counter += 1
                            secondary = SecondaryDimension(
                                name=col2_val,
                                code=self._generate_code(col2_val),
                                sort_order=secondary_counter,
                                metric_type=metric_type
                            )
                            current_primary.secondary_dimensions.append(secondary)
                        
            # 只有一级维度（没有层级名）
            elif col1_val and col1_val.strip():
                # 确保有当前层级
                if current_level is None:
                    level_counter += 1
                    current_level = LevelInfo(
                        name=f"默认层级{level_counter}",
                        description="",
                        sort_order=level_counter
                    )
                
                # 检查一级维度是否已存在
                existing_primary = None
                for p in current_level.primary_dimensions:
                    if p.name == col1_val:
                        existing_primary = p
                        break

                if existing_primary:
                    current_primary = existing_primary
                else:
                    primary_counter += 1
                    current_primary = PrimaryDimension(
                        name=col1_val,
                        code=self._generate_code(col1_val),
                        sort_order=primary_counter
                    )
                    current_level.primary_dimensions.append(current_primary)

                # 如果有二级维度
                if col2_val and col2_val.strip():
                    # 解析指标性质
                    metric_type = self._parse_metric_type(col3_val)
                    # 检查二级维度是否已存在
                    existing_sec = None
                    for s in current_primary.secondary_dimensions:
                        if s.name == col2_val:
                            existing_sec = s
                            break
                    if existing_sec:
                        continue  # 跳过重复的二级维度
                    secondary_counter += 1
                    secondary = SecondaryDimension(
                        name=col2_val,
                        code=self._generate_code(col2_val),
                        sort_order=secondary_counter,
                        metric_type=metric_type
                    )
                    current_primary.secondary_dimensions.append(secondary)
                    
            # 只有二级维度（继承当前的一级维度）
            elif col2_val and col2_val.strip() and current_primary:
                # 检查二级维度是否已存在
                existing_sec = None
                for s in current_primary.secondary_dimensions:
                    if s.name == col2_val:
                        existing_sec = s
                        break
                if existing_sec:
                    continue  # 跳过重复的二级维度
                secondary_counter += 1
                # 解析指标性质
                metric_type = self._parse_metric_type(col3_val)
                secondary = SecondaryDimension(
                    name=col2_val,
                    code=self._generate_code(col2_val),
                    sort_order=secondary_counter,
                    metric_type=metric_type
                )
                current_primary.secondary_dimensions.append(secondary)
        
        # 保存最后一个层级
        if current_level:
            self.levels.append(current_level)
        
        # 计算统计信息
        statistics = self._calculate_statistics()
        self.statistics = statistics  # Store for to_dict()

        return ParsedResult(
            levels=self.levels,
            statistics=statistics
        )
    
    def _get_cell_value(self, df: pd.DataFrame, row_idx: int, col_idx: int) -> str:
        """获取单元格值"""
        try:
            val = df.iloc[row_idx, col_idx]
            if pd.isna(val):
                return ""
            return str(val).strip()
        except:
            return ""
    
    def _is_level_row(self, df: pd.DataFrame, col0_val: str, row_idx: int) -> bool:
        """
        判断是否是层级标题行

        策略：
        1. 检查是否是合并单元格
        2. 检查是否包含描述特征（对应级别、具体职级、评估重点）
        3. 检查列2是否为空（层级行通常没有一级维度）
        """
        if not col0_val:
            return False

        # 检查是否是合并单元格
        for (min_row, max_row), value in self.merged_cells.items():
            if min_row <= row_idx + 1 <= max_row:  # Excel行号从1开始
                return True

        # 检查描述特征
        description_keywords = ['对应级别', '具体职级', '评估重点', '对应级别：']
        if any(keyword in col0_val for keyword in description_keywords):
            return True

        # 检查是否只有列0有值，列1和列2为空（在新行开始时）
        # 这种情况下，第一列是很长的文本（描述）
        if len(col0_val) > 50 and not self._has_content_in_cols(df, row_idx, [1, 2]):
            return True

        return False
    
    def _has_content_in_cols(self, df: pd.DataFrame, row_idx: int, col_indices: List[int]) -> bool:
        """检查指定行在指定列是否有内容"""
        try:
            for col_idx in col_indices:
                val = df.iloc[row_idx, col_idx]
                if not pd.isna(val) and str(val).strip():
                    return True
            return False
        except:
            return False
    
    def _extract_level_name(self, text: str) -> str:
        """从层级文本中提取名称"""
        if not text:
            return "未命名层级"
        
        # 描述文本通常以层级名称开头，后面是换行和描述
        # 例如："战略谋划\n对应级别：..."
        lines = text.split('\n')
        if lines:
            first_line = lines[0].strip()
            if first_line:
                return first_line
        
        # 如果没有换行，取前20个字符
        return text[:20].strip()
    
    def _extract_level_description(self, text: str) -> str:
        """从层级文本中提取描述"""
        if not text or len(text) < 50:
            return ""
        
        # 描述通常是对应级别、具体职级、评估重点部分
        # 直接返回整个文本作为描述
        return text
    
    def _generate_code(self, name: str) -> str:
        """生成编码"""
        if not name:
            return "unnamed"
        
        # 移除非字母数字字符（保留中文）
        code = re.sub(r'[^\w\u4e00-\u9fff]', '_', name)
        # 多个下划线合并为一个
        code = re.sub(r'_+', '_', code)
        # 转小写
        code = code.lower()
        # 移除首尾下划线
        code = code.strip('_')
        
        return code if code else "unnamed"
    
    def _parse_metric_type(self, value: str) -> str:
        """
        解析指标性质
        
        Args:
            value: 单元格值，如 "定性", "定量"
            
        Returns:
            'QUALITATIVE' (定性) 或 'QUANTITATIVE' (定量)
        """
        if not value:
            return 'QUALITATIVE'  # 默认定性
        
        value = value.strip().upper()
        
        if '定性' in value or value == 'QUALITATIVE':
            return 'QUALITATIVE'
        elif '定量' in value or value == 'QUANTITATIVE':
            return 'QUANTITATIVE'
        else:
            return 'QUALITATIVE'  # 默认定性
    
    def _calculate_statistics(self) -> Dict:
        """计算统计信息"""
        total_primary = sum(len(l.primary_dimensions) for l in self.levels)
        total_secondary = sum(
            sum(len(p.secondary_dimensions) for p in l.primary_dimensions)
            for l in self.levels
        )
        
        return {
            'levelCount': len(self.levels),
            'totalPrimaryDimensions': total_primary,
            'totalSecondaryDimensions': total_secondary,
            'levelNames': [l.name for l in self.levels],
            'primaryDimensionCountPerLevel': {
                l.name: len(l.primary_dimensions) for l in self.levels
            }
        }
    
    def to_dict(self) -> Dict:
        """转换为字典格式（API响应）"""
        return {
            'levels': [
                {
                    'name': level.name,
                    'description': level.description,
                    'sortOrder': level.sort_order,
                    'primaryDimensions': [
                        {
                            'name': primary.name,
                            'code': primary.code,
                            'sortOrder': primary.sort_order,
                            'secondaryDimensions': [
                                {
                                    'name': sec.name,
                                    'code': sec.code,
                                    'sortOrder': sec.sort_order,
                                    'metricType': sec.metric_type  # 指标性质
                                }
                                for sec in primary.secondary_dimensions
                            ]
                        }
                        for primary in level.primary_dimensions
                    ]
                }
                for level in self.levels
            ],
            'statistics': self.statistics
        }
    
    def to_json(self) -> str:
        """转换为JSON字符串"""
        return json.dumps(self.to_dict(), ensure_ascii=False, indent=2)

# ============================================================================
# 辅助函数
# ============================================================================

def parse_ahp_matrix_input(priorities: Dict[str, int]) -> Tuple[List[str], List[List[float]]]:
    """
    根据优先级构建AHP判断矩阵
    
    Args:
        priorities: {维度名: 优先级}，数字越小优先级越高
        
    Returns:
        (维度名列表, 判断矩阵)
    """
    if not priorities:
        return [], []
    
    dim_names = sorted(priorities.keys(), key=lambda x: priorities[x])
    n = len(dim_names)
    
    # 构建判断矩阵
    matrix = [[1.0 if i == j else 0.0 for j in range(n)] for i in range(n)]
    
    for i in range(n):
        for j in range(n):
            if i == j:
                matrix[i][j] = 1.0
            else:
                priority_i = priorities[dim_names[i]]
                priority_j = priorities[dim_names[j]]
                
                diff = abs(priority_i - priority_j)
                
                if priority_i < priority_j:
                    # i 的优先级高于 j
                    matrix[i][j] = diff + 1
                else:
                    # j 的优先级高于 i
                    matrix[i][j] = 1.0 / (diff + 1)
    
    return dim_names, matrix


def calculate_ahp_weights(matrix: List[List[float]]) -> Tuple[List[float], float]:
    """
    AHP计算权重和一致性比率
    
    Args:
        matrix: 判断矩阵
        
    Returns:
        (权重列表, 一致性比率CR)
    """
    n = len(matrix)
    
    # 转换为numpy数组
    mat = np.array(matrix, dtype=float)
    
    # 计算特征值和特征向量
    eigenvalues, eigenvectors = np.linalg.eig(mat)
    
    # 找最大特征值
    max_eigenvalue = float(np.max(eigenvalues.real))
    
    # 归一化特征向量作为权重
    max_eigenvector = eigenvectors[:, np.argmax(eigenvalues.real)].real
    weights = max_eigenvector / max_eigenvector.sum()
    
    # 计算一致性比率
    CI = (max_eigenvalue - n) / (n - 1) if n > 1 else 0
    
    RI_dict = {
        1: 0, 2: 0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24,
        7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49
    }
    RI = RI_dict.get(n, 1.41)
    CR = CI / RI if RI > 0 else 0
    
    return weights.tolist(), CR


# ============================================================================
# 主入口
# ============================================================================

if __name__ == '__main__':
    try:
        input_json = sys.stdin.read().strip()
        
        if input_json:
            input_data = json.loads(input_json)
            action = input_data.get('action', 'parse')
            
            if action == 'parse':
                # 解析Excel
                file_path = input_data.get('filePath')
                
                if file_path:
                    parser = DynamicIndicatorParser()
                    result = parser.parse_file(file_path)
                    print(parser.to_json())
                else:
                    print(json.dumps({
                        'success': False,
                        'message': '缺少文件路径'
                    }, ensure_ascii=False))
                    
            elif action == 'ahp':
                # 计算AHP权重
                priorities = input_data.get('priorities', {})
                dim_names, matrix = parse_ahp_matrix_input(priorities)
                weights, cr = calculate_ahp_weights(matrix)
                
                result = {
                    'success': True,
                    'dimensionNames': dim_names,
                    'judgmentMatrix': matrix,
                    'weights': {name: weights[i] for i, name in enumerate(dim_names)},
                    'consistencyRatio': cr,
                    'consistencyPassed': cr < 0.1
                }
                print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            print(json.dumps({
                'success': False,
                'message': '无输入数据'
            }, ensure_ascii=False))
            
    except Exception as e:
        import traceback
        error_result = {
            'success': False,
            'message': f'执行失败: {str(e)}',
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        print(json.dumps(error_result, ensure_ascii=False, indent=2))
        sys.exit(1)
